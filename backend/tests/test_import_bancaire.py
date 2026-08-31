import io
import unicodedata
from datetime import date, datetime

import openpyxl
import pytest

from app import crud, models, schemas
from app.constants import (
    COLONNES_IMPORT_PAR_DEFAUT,
    DomaineImport,
    ModeComparaison,
    Sens,
    Statut,
)
from app.services import import_bancaire

from .conftest import creer_compte, get_categorie_id, get_monnaie_id, get_type_id


def _construire_fichier(lignes: list[dict]) -> bytes:
    """Construit un classeur à 12 colonnes façon export bancaire. `lignes` ne
    précise que les colonnes utilisées par le parseur (date/nature/categorie/
    montant/compte) ; les autres restent vides, comme dans un vrai relevé."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.append([f"Colonne {i}" for i in range(1, 13)])  # en-tête, ignorée
    for ligne in lignes:
        row = [None] * 12
        row[0] = ligne.get("date")
        row[3] = ligne.get("nature")
        row[5] = ligne.get("categorie")
        row[6] = ligne.get("montant")
        row[9] = ligne.get("compte")
        feuille.append(row)
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def _construire_fichier_avec_reference(lignes: list[dict]) -> bytes:
    """Comme _construire_fichier, avec en plus une colonne 12 ("Référence")
    non mappée par la configuration par défaut : sert uniquement aux tests de
    détection de doublons (colonnes exclues de la comparaison)."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.append([f"Colonne {i}" for i in range(1, 13)])
    for ligne in lignes:
        row = [None] * 12
        row[0] = ligne.get("date")
        row[3] = ligne.get("nature")
        row[5] = ligne.get("categorie")
        row[6] = ligne.get("montant")
        row[9] = ligne.get("compte")
        row[11] = ligne.get("reference")
        feuille.append(row)
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def _make_compte(db, nom="CC Perso"):
    return creer_compte(db, nom)


def _make_preset(db, nom="Défaut", colonnes=None, colonnes_exclues=None, ignorer_premiere_ligne=True):
    """ignorer_premiere_ligne=True par défaut ici parce que les deux
    constructeurs de fichier ci-dessus écrivent une ligne d'en-tête. Le défaut
    applicatif est l'inverse (False : la première ligne est une donnée) —
    couvert par les tests dédiés en fin de fichier."""
    return crud.create_import_preset(
        db,
        nom,
        colonnes if colonnes is not None else COLONNES_IMPORT_PAR_DEFAUT,
        colonnes_exclues or [],
        ignorer_premiere_ligne=ignorer_premiere_ligne,
    )


def test_lire_lignes_brutes_extrait_les_bonnes_colonnes(db_session):
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses Monoprix",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )

    lignes = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True
    )

    assert len(lignes) == 1
    assert lignes[0]["nature"] == "Courses Monoprix"
    assert lignes[0]["categorie_banque"] == "Alimentation"
    assert lignes[0]["montant_brut"] == -45.2
    assert lignes[0]["compte_banque"] == "CC Perso"


def test_lire_lignes_brutes_capture_toutes_les_colonnes(db_session):
    # Colonne 12 ("Référence") n'est mappée par aucune propriété de l'app :
    # elle doit quand même apparaître dans donnees_completes.
    contenu = _construire_fichier_avec_reference(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses Monoprix",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
                "reference": "REF-001",
            }
        ]
    )

    lignes = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True
    )

    donnees = lignes[0]["donnees_completes"]
    # openpyxl relit toute cellule de date comme un datetime (Excel ne
    # distingue pas date/datetime en interne) -- cf. parser_date.
    assert donnees["1"] == datetime(2026, 7, 1).isoformat()
    assert donnees["4"] == "Courses Monoprix"
    assert donnees["6"] == "Alimentation"
    assert donnees["7"] == -45.2
    assert donnees["10"] == "CC Perso"
    assert donnees["12"] == "REF-001"
    # Colonnes vides (2, 3, 5, 8, 9, 11) : jamais stockées.
    assert "2" not in donnees


@pytest.mark.parametrize(
    "valeur",
    [
        "14/07/2026",
        "14/07/2026 09:32",
        "14/07/2026 09:32:07",
        "2026-07-14 09:32",
        "2026-07-14T09:32:07",
        "14-07-2026 09:32",
        datetime(2026, 7, 14, 9, 32),
    ],
)
def test_une_date_horodatee_est_lue_sans_son_heure(valeur):
    """Les néobanques datent à la minute ; une opération, elle, est du jour.
    L'heure est retirée dès la lecture — l'app ne stocke qu'une date, et tout
    ce qui s'affiche est donc au même format partout."""
    assert import_bancaire.parser_date(valeur) == date(2026, 7, 14)


def test_lapercu_du_fichier_affiche_les_dates_sans_leur_heure(db_session):
    """« Le fichier tel qu'il est » montrait « 2026-07-14T09:32:00 » là où
    l'aperçu des lignes, deux blocs plus bas, écrivait « 14/07/2026 » pour la
    même ligne."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": datetime(2026, 7, 14, 9, 32),
                "nature": "Courses",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )

    apercu = import_bancaire.previsualiser(
        db_session, preset.id, contenu, compte_id_defaut=compte.id
    ).apercu_fichier

    # Ligne 1 = l'en-tête écrit par _construire_fichier, ligne 2 = les données.
    assert apercu.lignes[1][0] == "14/07/2026"


def test_lapercu_du_fichier_ne_reformate_pas_ce_qui_nest_pas_une_date(db_session):
    """Une référence bancaire peut ressembler à une date sans en être une :
    la retoucher ferait mentir une vue qui promet de montrer le fichier tel
    qu'il est."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier_avec_reference(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
                "reference": "20260714 BP0019471",
            }
        ]
    )

    apercu = import_bancaire.previsualiser(
        db_session, preset.id, contenu, compte_id_defaut=compte.id
    ).apercu_fichier

    assert apercu.lignes[1][11] == "20260714 BP0019471"


def test_lire_lignes_brutes_lit_un_csv_avec_point_virgule(db_session):
    # Même format à 12 colonnes que _construire_fichier, mais en CSV ";"
    # (standard des exports bancaires français) plutôt qu'en xlsx.
    contenu = (
        ";".join(f"Colonne {i}" for i in range(1, 13)) + "\r\n"
        "01/07/2026;;;Courses Monoprix;;Alimentation;-45,2;;;CC Perso;;\r\n"
    ).encode("utf-8")

    lignes = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True
    )

    assert len(lignes) == 1
    assert lignes[0]["nature"] == "Courses Monoprix"
    assert lignes[0]["categorie_banque"] == "Alimentation"
    assert lignes[0]["montant_brut"] == "-45,2"
    assert lignes[0]["compte_banque"] == "CC Perso"
    # parser_montant tolère la virgule décimale française une fois résolu.
    assert import_bancaire.parser_montant(lignes[0]["montant_brut"]) == -45.2


def test_lire_lignes_brutes_decode_un_csv_cp1252_avec_accents(db_session):
    texte = (
        ";".join(f"Colonne {i}" for i in range(1, 13)) + "\r\n"
        "01/07/2026;;;Café Déjeuner;;Alimentation;-12,5;;;CC Perso;;\r\n"
    )
    contenu = texte.encode("cp1252")

    lignes = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True
    )

    assert lignes[0]["nature"] == "Café Déjeuner"


def test_previsualiser_fonctionne_a_partir_dun_csv(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    crud.set_mapping_categorie(db_session, preset.id, "Alimentation", categorie_id)
    crud.set_mapping_compte(db_session, preset.id, "CC Perso", compte.id)

    contenu = (
        ";".join(f"Colonne {i}" for i in range(1, 13)) + "\r\n"
        "01/07/2026;;;Courses;;Alimentation;-45,2;;;CC Perso;;\r\n"
    ).encode("utf-8")

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    assert len(preview.lignes) == 1
    assert preview.lignes[0].categorie_id == categorie_id
    assert preview.lignes[0].compte_id == compte.id
    assert preview.lignes[0].montant == 45.2
    assert preview.lignes[0].erreur is None


def test_previsualiser_resout_les_mappings_connus(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    crud.set_mapping_categorie(db_session, preset.id, "Alimentation", categorie_id)
    crud.set_mapping_compte(db_session, preset.id, "CC Perso", compte.id)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    assert len(preview.lignes) == 1
    ligne = preview.lignes[0]
    assert ligne.categorie_id == categorie_id
    assert ligne.compte_id == compte.id
    # Toujours positif, peu importe le signe dans le fichier (sortie = négatif côté banque).
    assert ligne.montant == 45.2
    # Le signe d'origine reste disponible séparément (utile pour distinguer
    # émetteur/récepteur d'un virement interne, côté frontend).
    assert ligne.montant_signe == -45.2
    assert ligne.erreur is None
    assert preview.categories_inconnues == []
    assert preview.comptes_inconnus == []


def test_previsualiser_liste_les_categories_et_comptes_inconnus(db_session):
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    # "Alimentation" reste signalée comme à confirmer même si une valeur par
    # défaut (Autres) lui a déjà été proposée ; le compte, lui, n'a pas de
    # valeur par défaut sensée et reste donc bloquant tant qu'il n'est pas mappé.
    assert preview.categories_inconnues == ["Alimentation"]
    assert preview.comptes_inconnus == ["CC Perso"]
    assert preview.lignes[0].categorie_id == get_categorie_id(db_session, "Autres")
    assert preview.lignes[0].categorie_suggestion_auto is True
    assert preview.lignes[0].compte_id is None


def test_previsualiser_ne_suggere_jamais_virement_interne_automatiquement(db_session):
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Virement vers autre compte",
                "montant": -200.0,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    # Comme toute autre catégorie non mappée, elle est proposée dans "Autres" ;
    # "Virement interne" n'est jamais déduit automatiquement du nom bancaire.
    ligne = preview.lignes[0]
    assert ligne.categorie_id == get_categorie_id(db_session, "Autres")
    assert ligne.categorie_suggestion_auto is True
    assert ligne.montant == 200.0
    assert ligne.montant_signe == -200.0


def test_confirmer_categorie_inconnue_est_creee_dans_autres_sans_bloquer(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    # Aucun override de catégorie fourni : elle doit quand même s'importer,
    # classée automatiquement dans "Autres".
    overrides = schemas.ImportMappingOverrides(comptes={"CC Perso": compte.id})

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    assert resultat.lignes_ignorees == []
    operation = db_session.query(models.Operation).one()
    assert operation.categorie_id == get_categorie_id(db_session, "Autres")


def test_confirmer_cree_une_operation_classique(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    assert resultat.lignes_ignorees == []

    operation = db_session.query(models.Operation).one()
    assert operation.nature == "Courses"
    assert operation.montant == 45.2
    assert operation.sens == Sens.depense
    assert operation.statut == Statut.reel
    assert operation.remboursable is False
    assert operation.compte_id == compte.id
    assert operation.categorie_id == categorie_id


def test_confirmer_persiste_les_mappings_pour_le_prochain_import(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    # Un import ultérieur, sans overrides, résout automatiquement grâce au
    # mapping mémorisé lors du premier import.
    contenu2 = _construire_fichier(
        [
            {
                "date": date(2026, 7, 2),
                "nature": "Boulangerie",
                "categorie": "Alimentation",
                "montant": -5.0,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu2)

    assert preview.lignes[0].categorie_id == categorie_id
    assert preview.lignes[0].compte_id == compte.id
    assert preview.categories_inconnues == []
    assert preview.comptes_inconnus == []


def test_confirmer_ignore_les_lignes_sans_compte_resolu(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            },
            {
                "date": date(2026, 7, 2),
                "nature": "Essence",
                "categorie": "Alimentation",
                "montant": -60.0,
                "compte": "Compte inconnu",
            },
        ]
    )
    # Seul "CC Perso" est mappé ; "Compte inconnu" reste non résolu (pas de
    # valeur par défaut sensée pour un compte, contrairement à une catégorie).
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    assert len(resultat.lignes_ignorees) == 1
    assert resultat.lignes_ignorees[0].nature == "Essence"
    assert "compte" in resultat.lignes_ignorees[0].erreur


def test_confirmer_applique_une_modification_manuelle_de_ligne(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    autre_categorie_id = get_categorie_id(db_session, "Loisirs & sorties")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        # La ligne 2 (après l'en-tête) est corrigée à la main : nature, montant
        # et catégorie diffèrent de ce que le fichier/mapping auraient résolu.
        lignes={2: schemas.ImportLigneOverride(nature="Restaurant", montant=30.0, categorie_id=autre_categorie_id)},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.nature == "Restaurant"
    assert operation.montant == 30.0
    assert operation.categorie_id == autre_categorie_id


def test_confirmer_applique_un_montant_du_manuel(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Facture partagee",
                "categorie": "Alimentation",
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="remboursable", montant_du=40.0)},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    # Une opération importée est toujours réelle (transaction déjà survenue).
    assert operation.statut == Statut.reel
    assert operation.montant_du == 40.0
    assert operation.montant_a_rembourser == 40.0


def test_confirmer_clampe_un_montant_du_manuel_superieur_au_montant(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -20.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="remboursable", montant_du=999.0)},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.montant_du == 20.0
    assert operation.montant_a_rembourser == 20.0


def test_confirmer_une_modification_manuelle_peut_corriger_une_ligne_illisible(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": "n'importe quoi",
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(date=date(2026, 7, 3))},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    assert resultat.lignes_ignorees == []
    operation = db_session.query(models.Operation).one()
    assert operation.date == date(2026, 7, 3)


def test_confirmer_reclasse_une_ligne_en_depense_remboursable(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="remboursable")},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.remboursable is True
    assert operation.montant_du == 45.2
    assert operation.montant_a_rembourser == 45.2


def test_confirmer_reclasse_une_ligne_en_pret_force_remboursable(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Prêt à un ami",
                "categorie": "Divers",
                "montant": 100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte.id},
        # Le type porte à lui seul le caractère remboursable : nul besoin de
        # le préciser à la ligne.
        lignes={2: schemas.ImportLigneOverride(type_code="pret")},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.type_code == "pret"
    # Un type à catégorie imposée ne porte plus de catégorie : le type est sa
    # classification.
    assert operation.categorie_id is None
    assert operation.remboursable is True
    assert operation.montant_du == 100.0
    assert operation.montant_a_rembourser == 100.0


def test_confirmer_reclasse_une_ligne_en_remboursement_jamais_remboursable(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Remboursement ami",
                "categorie": "Divers",
                "montant": 20.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte.id},
        # Un type de règlement n'est jamais remboursable, quoi qu'on demande.
        lignes={2: schemas.ImportLigneOverride(type_code="remboursements")},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.type_code == "remboursements"
    assert operation.categorie_id is None
    assert operation.remboursable is False
    assert operation.montant_du == 0.0
    assert operation.montant_a_rembourser == 0.0


def test_confirmer_refuse_un_virement_sans_second_compte(db_session):
    """Un virement interne décrit DEUX comptes. N'en importer qu'un laissait une
    écriture orpheline, qu'il fallait retrouver et compléter à la main plus
    tard : la ligne est désormais refusée tant que l'autre côté n'est pas
    désigné."""
    compte = _make_compte(db_session, "CC Perso")
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="virement")},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 0
    assert db_session.query(models.Operation).count() == 0
    assert "compte en face" in resultat.lignes_ignorees[0].erreur


def test_confirmer_cree_le_virement_des_que_le_second_compte_est_designe(db_session):
    """La contrepartie du test précédent : avec les deux comptes, c'est un vrai
    virement double-écriture, orienté par le signe du montant bancaire."""
    source = _make_compte(db_session, "CC Perso")
    destination = _make_compte(db_session, "Livret A")
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": source.id},
        lignes={
            2: schemas.ImportLigneOverride(
                type_code="virement", compte_id_autre=destination.id
            )
        },
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 2
    sortante = db_session.query(models.Operation).filter_by(sens=Sens.transfert_sortant).one()
    entrante = db_session.query(models.Operation).filter_by(sens=Sens.transfert_entrant).one()
    assert sortante.compte_id == source.id
    assert entrante.compte_id == destination.id
    assert sortante.virement_id == entrante.virement_id


def test_confirmer_virement_categorie_sans_aucun_compte_est_ignore(db_session):
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                "montant": -100.0,
                "compte": "Compte inconnu",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        lignes={2: schemas.ImportLigneOverride(type_code="virement")},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 0
    assert len(resultat.lignes_ignorees) == 1
    assert "compte non résolu" in resultat.lignes_ignorees[0].erreur
    assert db_session.query(models.Operation).count() == 0


def test_confirmer_virement_avec_les_deux_comptes_cree_un_vrai_virement(db_session):
    compte_source = _make_compte(db_session, "CC Perso")
    compte_destination = _make_compte(db_session, "Livret A")
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                # Négatif : "CC Perso" (compte connu du fichier) est l'émetteur.
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte_source.id},
        lignes={
            2: schemas.ImportLigneOverride(type_code="virement", compte_id_autre=compte_destination.id)
        },
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 2
    operations = db_session.query(models.Operation).all()
    assert len(operations) == 2
    assert operations[0].virement_id == operations[1].virement_id
    sortante = next(o for o in operations if o.sens == Sens.transfert_sortant)
    entrante = next(o for o in operations if o.sens == Sens.transfert_entrant)
    assert sortante.compte_id == compte_source.id
    assert entrante.compte_id == compte_destination.id
    assert sortante.montant == 100.0
    assert entrante.montant == 100.0


def test_confirmer_virement_meme_compte_des_deux_cotes_est_ignore(db_session):
    compte = _make_compte(db_session, "CC Perso")
    preset = _make_preset(db_session)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="virement", compte_id_autre=compte.id)},
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 0
    assert len(resultat.lignes_ignorees) == 1
    assert "différents" in resultat.lignes_ignorees[0].erreur
    assert db_session.query(models.Operation).count() == 0


def test_confirmer_ignore_une_ligne_supprimee_manuellement(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            },
            {
                "date": date(2026, 7, 2),
                "nature": "Essence",
                "categorie": "Alimentation",
                "montant": -60.0,
                "compte": "CC Perso",
            },
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes_supprimees=[3],
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    assert resultat.lignes_ignorees == []
    operation = db_session.query(models.Operation).one()
    assert operation.nature == "Courses"


def test_confirmer_avec_compte_id_defaut_quand_pas_de_colonne_compte(db_session):
    compte = _make_compte(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    colonnes_sans_compte = [c for c in COLONNES_IMPORT_PAR_DEFAUT if c["propriete"] != "compte_banque"]
    preset = _make_preset(db_session, colonnes=colonnes_sans_compte)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(categories={"Alimentation": categorie_id})

    resultat = import_bancaire.confirmer(
        db_session, preset.id, contenu, overrides, compte_id_defaut=compte.id
    )

    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.compte_id == compte.id


def test_confirmer_ignore_les_lignes_avec_donnees_illisibles(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")

    contenu = _construire_fichier(
        [
            {
                "date": "n'importe quoi",
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 0
    assert len(resultat.lignes_ignorees) == 1
    assert "date" in resultat.lignes_ignorees[0].erreur


def test_confirmer_enregistre_un_historique(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )

    import_bancaire.confirmer(db_session, preset.id, contenu, overrides, nom_fichier="releve_juillet.xlsx")

    historique = crud.get_import_historique(db_session, preset.id)
    assert len(historique) == 1
    assert historique[0].nom_fichier == "releve_juillet.xlsx"
    assert historique[0].operations_creees == 1
    assert historique[0].lignes_ignorees == 0


def test_confirmer_puis_reimporter_le_meme_fichier_detecte_un_doublon(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    # Réimporter exactement le même fichier (même preset) : la ligne reste
    # une ImportLigne normale (même format), mais porte doublon_de, et la
    # ligne existante correspondante est résolue au même format pour comparaison.
    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    assert len(preview.lignes) == 1
    ligne_doublon = preview.lignes[0]
    assert ligne_doublon.doublon_de is not None
    ligne_existante = preview.lignes_existantes[str(ligne_doublon.doublon_de)]
    assert ligne_existante.nature == ligne_doublon.nature
    assert ligne_existante.montant == ligne_doublon.montant

    # Le doublon est signalé mais reste importable : c'est le frontend qui le
    # pré-sélectionne et bloque la confirmation tant qu'il n'a pas été traité.
    # Laissé passer ici (aucune ligne supprimée), il crée bien son opération.
    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)
    assert resultat.operations_creees == 1
    assert resultat.doublons_detectes == 1
    assert db_session.query(models.Operation).count() == 2
    assert db_session.query(models.LigneImportBrute).count() == 2


def test_confirmer_un_doublon_supprime_ne_cree_pas_dopration(db_session):
    # Le chemin normal : l'utilisateur supprime les doublons pré-sélectionnés
    # dans l'aperçu avant de confirmer.
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    doublons = [l.ligne for l in preview.lignes if l.doublon_de is not None]
    assert doublons

    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        overrides.model_copy(update={"lignes_supprimees": doublons}),
    )

    assert resultat.operations_creees == 0
    assert resultat.doublons_detectes == 1
    assert db_session.query(models.Operation).count() == 1
    assert db_session.query(models.LigneImportBrute).count() == 1


def test_supprimer_une_operation_la_retire_du_stock_anti_doublons(db_session):
    """Le lien LigneImportBrute.operation_id rend le relevé réimportable après
    suppression : sans lui, la ligne restait au stock et la réimportation
    était éternellement bloquée en doublon."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    operation = db_session.query(models.Operation).one()
    ligne_brute = db_session.query(models.LigneImportBrute).one()
    assert ligne_brute.operation_id == operation.id

    crud.delete_operation(db_session, operation)

    assert db_session.query(models.LigneImportBrute).count() == 0
    # Le même fichier n'est plus vu comme un doublon.
    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    assert preview.lignes[0].doublon_de is None


def test_colonne_exclue_de_la_comparaison_permet_de_detecter_un_doublon_malgre_une_difference(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )

    premier_fichier = _construire_fichier_avec_reference(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
                "reference": "REF-001",
            }
        ]
    )
    import_bancaire.confirmer(db_session, preset.id, premier_fichier, overrides)

    # Même opération, mais la colonne 12 (référence bancaire, propre à
    # l'export) diffère -- comme un même relevé réexporté un autre jour.
    second_fichier = _construire_fichier_avec_reference(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
                "reference": "REF-002",
            }
        ]
    )

    # Sans exclusion configurée : la différence en colonne 12 empêche la
    # détection (comparaison stricte sur toutes les colonnes).
    preview = import_bancaire.previsualiser(db_session, preset.id, second_fichier)
    assert len(preview.lignes) == 1
    assert preview.lignes[0].doublon_de is None

    # Colonne 12 exclue de la comparaison : la même ligne est maintenant
    # détectée comme un doublon malgré la différence de référence.
    crud.update_import_preset(
        db_session, preset, colonnes=COLONNES_IMPORT_PAR_DEFAUT, colonnes_comparaison=[12]
    )
    preview = import_bancaire.previsualiser(db_session, preset.id, second_fichier)
    assert len(preview.lignes) == 1
    assert preview.lignes[0].doublon_de is not None


def test_confirmer_ne_stocke_que_les_lignes_devenues_des_operations(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            },
            {
                "date": date(2026, 7, 2),
                "nature": "Essence",
                "categorie": "Alimentation",
                "montant": -60.0,
                "compte": "CC Perso",
            },
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id},
        comptes={"CC Perso": compte.id},
        lignes_supprimees=[3],
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 1
    # Seule la ligne réellement importée rejoint le stock : celle supprimée à
    # la main n'a rien créé, la revoir au prochain import est voulu (sinon
    # elle serait signalée en doublon d'une opération inexistante).
    ligne_brute = db_session.query(models.LigneImportBrute).one()
    operation = db_session.query(models.Operation).one()
    assert ligne_brute.operation_id == operation.id


def test_doublon_non_detecte_entre_deux_presets_differents(db_session):
    compte = _make_compte(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    preset_a = _make_preset(db_session, nom="Banque A")
    preset_b = _make_preset(db_session, nom="Banque B")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(db_session, preset_a.id, contenu, overrides)

    # Même fichier, mais importé sous un preset différent : le stock de
    # doublons est isolé par preset, donc pas de doublon détecté ici.
    preview = import_bancaire.previsualiser(db_session, preset_b.id, contenu)
    assert preview.lignes[0].doublon_de is None

    resultat = import_bancaire.confirmer(db_session, preset_b.id, contenu, overrides)
    assert resultat.operations_creees == 1
    assert resultat.doublons_detectes == 0
    assert db_session.query(models.Operation).count() == 2


def _regle_virement(db):
    """Seule une règle peut poser le type d'une ligne importée."""
    return crud.create_regle_categorisation(
        db,
        nom="Transferts",
        type_id=get_type_id(db, "virement"),
        categorie_id=None,
        conditions={
            "operateur": "ET",
            "groupes": [
                {
                    "operateur": "ET",
                    "conditions": [
                        {"champ": "nature", "operateur": "contient", "valeur": "Transfert"}
                    ],
                }
            ],
        },
    )


def test_virement_sortant_sans_devise_place_le_montant_du_cote_envoye(db_session):
    """Un relevé ordinaire (ni colonne de devise, ni colonne « Montant
    initial ») n'écrit qu'un montant, dans la monnaie de SON compte. Sur une
    ligne sortante, c'est ce qui PART.

    Le placer dans `montant` — qui décrit ce qui ARRIVE depuis la migration
    0029 — inversait les deux jambes : l'aperçu montrait le débit en « montant
    reçu » et laissait « montant initial (envoyé) » vide."""
    source = _make_compte(db_session, nom="CC Perso")
    destination = _make_compte(db_session, nom="Livret")
    _regle_virement(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Transfert vers Livret",
                "montant": -300.0,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    ligne = preview.lignes[0]
    assert ligne.montant_envoye == 300.0, "ce qui part doit porter le montant lu"
    assert ligne.montant is None, "ce qui arrive est inconnu tant qu'aucune devise ne le dit"
    assert ligne.montant_envoye_deduit is True
    # La ligne n'est PAS en erreur pour autant : il ne lui manque que le compte
    # d'en face, que l'utilisateur désigne dans l'aperçu.
    assert "montant illisible" not in (ligne.erreur or "")

    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        schemas.ImportMappingOverrides(
            comptes={"CC Perso": source.id},
            lignes={2: schemas.ImportLigneOverride(compte_id_autre=destination.id)},
        ),
    )

    # Sans change, la jambe manquante vaut l'autre : les deux écritures portent
    # 300 €, dans le bon sens.
    assert resultat.operations_creees == 2
    sortante = db_session.query(models.Operation).filter_by(sens="transfert_sortant").one()
    entrante = db_session.query(models.Operation).filter_by(sens="transfert_entrant").one()
    assert (sortante.montant, sortante.compte_id) == (300.0, source.id)
    assert (entrante.montant, entrante.compte_id) == (300.0, destination.id)


def test_virement_sortant_sans_devise_vers_un_compte_dune_autre_monnaie(db_session):
    """Même relevé, mais le compte d'en face ne porte pas l'euro : ce qui
    arrive est réellement inconnu, et l'app n'invente aucun taux."""
    euro = get_monnaie_id(db_session)
    dollar = crud.create_monnaie(db_session, "Dollar", "$").id
    source = creer_compte(db_session, "CC Perso", monnaies=[(euro, 1000.0)])
    destination = creer_compte(db_session, "Wise USD", monnaies=[(dollar, 0.0)])
    _regle_virement(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Transfert vers Wise",
                "montant": -300.0,
                "compte": "CC Perso",
            }
        ]
    )

    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        schemas.ImportMappingOverrides(
            comptes={"CC Perso": source.id},
            lignes={2: schemas.ImportLigneOverride(compte_id_autre=destination.id)},
        ),
    )
    assert resultat.operations_creees == 0
    assert "montant reçu" in resultat.lignes_ignorees[0].erreur

    # Le montant reçu saisi dans l'aperçu complète la ligne : 300 € partent,
    # 325 $ arrivent.
    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        schemas.ImportMappingOverrides(
            comptes={"CC Perso": source.id},
            lignes={
                2: schemas.ImportLigneOverride(compte_id_autre=destination.id, montant=325.0)
            },
        ),
    )
    assert resultat.operations_creees == 2
    sortante = db_session.query(models.Operation).filter_by(sens="transfert_sortant").one()
    entrante = db_session.query(models.Operation).filter_by(sens="transfert_entrant").one()
    assert (sortante.montant, sortante.monnaie_id) == (300.0, euro)
    assert (entrante.montant, entrante.monnaie_id) == (325.0, dollar)


def test_ligne_classique_reclassee_en_virement_reoriente_le_montant(db_session):
    """Le cas de la reprise à la main, sans aucune règle.

    Le relevé n'a ni colonne « Sens » (le signe suffit) ni colonne de devise :
    la ligne est donc lue en `classique`, et son montant se range dans
    `montant`. La reclasser en virement interne dans l'aperçu doit REJOUER le
    raisonnement sur le signe — sortante, le montant lu est ce qui PART — comme
    si le fichier l'avait donnée en virement dès le départ.

    Sans ce rejeu, la ligne arrivait au compte d'en face avec « montant » plein
    et « montant initial » vide : sur un compte d'une autre devise, l'app
    réclamait la jambe émettrice qu'elle avait pourtant déjà lue, et prenait le
    débit du relevé pour un montant reçu."""
    euro = get_monnaie_id(db_session)
    dollar = crud.create_monnaie(db_session, "Dollar", "$").id
    source = creer_compte(db_session, "CC Perso", monnaies=[(euro, 1000.0)])
    destination = creer_compte(db_session, "Wise USD", monnaies=[(dollar, 0.0)])
    # Pas de _regle_virement ici : c'est tout l'objet du test.
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "PAIEMENT WISE",
                "montant": -300.0,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    ligne = preview.lignes[0]
    # Telle que lue : une dépense classique ordinaire, montant du bon côté.
    assert ligne.type_code == "classique"
    assert (ligne.montant, ligne.montant_envoye) == (300.0, None)

    # Reclassée en virement, sans que rien d'autre ne soit saisi : le montant
    # passe du côté « envoyé », et ce qui arrive reste inconnu — deux devises,
    # aucun taux. La ligne est refusée, en nommant précisément ce qui manque.
    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        schemas.ImportMappingOverrides(
            comptes={"CC Perso": source.id},
            lignes={
                2: schemas.ImportLigneOverride(
                    type_code="virement", compte_id_autre=destination.id
                )
            },
        ),
    )
    assert resultat.operations_creees == 0
    assert "montant reçu" in resultat.lignes_ignorees[0].erreur

    # Le montant reçu complété, le virement se crée avec ses deux jambes :
    # 300 € partent du compte du relevé, 325 $ arrivent sur l'autre.
    resultat = import_bancaire.confirmer(
        db_session,
        preset.id,
        contenu,
        schemas.ImportMappingOverrides(
            comptes={"CC Perso": source.id},
            lignes={
                2: schemas.ImportLigneOverride(
                    type_code="virement",
                    compte_id_autre=destination.id,
                    montant=325.0,
                    montant_envoye=300.0,
                    monnaie_id=dollar,
                    monnaie_envoyee_id=euro,
                )
            },
        ),
    )
    assert resultat.operations_creees == 2
    sortante = db_session.query(models.Operation).filter_by(sens="transfert_sortant").one()
    entrante = db_session.query(models.Operation).filter_by(sens="transfert_entrant").one()
    assert (sortante.montant, sortante.monnaie_id, sortante.compte_id) == (300.0, euro, source.id)
    assert (entrante.montant, entrante.monnaie_id, entrante.compte_id) == (
        325.0,
        dollar,
        destination.id,
    )


def test_mode_selection_ne_compare_que_les_colonnes_choisies(db_session):
    """Le pendant de l'exclusion : au lieu de recenser les colonnes qui bougent
    d'un export à l'autre, on désigne les seules qui identifient une ligne."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    ligne = {
        "date": date(2026, 7, 1),
        "nature": "Courses",
        "categorie": "Alimentation",
        "montant": -45.2,
        "compte": "CC Perso",
    }
    import_bancaire.confirmer(
        db_session, preset.id, _construire_fichier_avec_reference([{**ligne, "reference": "A1"}]), overrides
    )

    # Référence différente : en exclusion (défaut, liste vide), tout est
    # comparé — donc pas de doublon.
    autre = _construire_fichier_avec_reference([{**ligne, "reference": "B2"}])
    preview = import_bancaire.previsualiser(db_session, preset.id, autre)
    assert preview.lignes[0].doublon_de is None

    # Colonnes 1 (date), 4 (nature) et 7 (montant) seules comparées : la
    # référence, comme tout le reste, cesse de compter.
    crud.update_import_preset(
        db_session,
        preset,
        colonnes_comparaison=[1, 4, 7],
        mode_comparaison=ModeComparaison.selection.value,
    )
    preview = import_bancaire.previsualiser(db_session, preset.id, autre)
    assert preview.lignes[0].doublon_de is not None

    # Une des colonnes retenues change : ce n'est plus la même ligne.
    montant_different = _construire_fichier_avec_reference(
        [{**ligne, "montant": -99.9, "reference": "B2"}]
    )
    preview = import_bancaire.previsualiser(db_session, preset.id, montant_different)
    assert preview.lignes[0].doublon_de is None


def test_mode_selection_sans_colonne_ne_declare_aucun_doublon(db_session):
    """Ne comparer AUCUNE colonne rendrait toute ligne identique à la première
    en stock. Le routeur interdit d'enregistrer ça ; si une base y arrive
    autrement, le détecteur doit se taire plutôt que tout confondre."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    import_bancaire.confirmer(
        db_session,
        preset.id,
        _construire_fichier(
            [{"date": date(2026, 7, 1), "nature": "Courses", "categorie": "Alimentation", "montant": -45.2, "compte": "CC Perso"}]
        ),
        overrides,
    )
    crud.update_import_preset(
        db_session, preset, colonnes_comparaison=[], mode_comparaison=ModeComparaison.selection.value
    )

    # Une ligne qui n'a rien à voir avec celle en stock.
    preview = import_bancaire.previsualiser(
        db_session,
        preset.id,
        _construire_fichier(
            [{"date": date(2026, 9, 9), "nature": "Essence", "categorie": "Alimentation", "montant": -80.0, "compte": "CC Perso"}]
        ),
    )
    assert preview.lignes[0].doublon_de is None


def test_doublon_detecte_malgre_des_differences_invisibles(db_session):
    """Espace insécable, forme Unicode décomposée, espaces de bord : trois
    façons d'écrire un libellé qui s'affiche à l'identique — y compris dans la
    base, quand on va y vérifier. Elles faisaient échouer la détection sur des
    lignes que l'utilisateur voyait pourtant comme les mêmes."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    base = {
        "date": date(2026, 7, 1),
        "categorie": "Alimentation",
        "montant": -45.2,
        "compte": "CC Perso",
    }
    import_bancaire.confirmer(
        db_session,
        preset.id,
        _construire_fichier([{**base, "nature": "CARTE CAFÉ CRÈME"}]),
        overrides,
    )

    variantes = {
        "espace insécable": "CARTE CAFÉ CRÈME",
        "accents décomposés (NFD)": unicodedata.normalize("NFD", "CARTE CAFÉ CRÈME"),
        "espaces de bord": "  CARTE CAFÉ CRÈME ",
        "caractère de largeur nulle": "CARTE CAFÉ​ CRÈME",
    }
    for libelle, nature in variantes.items():
        preview = import_bancaire.previsualiser(
            db_session, preset.id, _construire_fichier([{**base, "nature": nature}])
        )
        assert preview.lignes[0].doublon_de is not None, libelle

    # La normalisation reste limitée à l'invisible : une différence qui SE VOIT
    # reste une ligne différente.
    preview = import_bancaire.previsualiser(
        db_session, preset.id, _construire_fichier([{**base, "nature": "CARTE CAFE CREME"}])
    )
    assert preview.lignes[0].doublon_de is None


def test_normaliser_pour_comparaison_unifie_nombre_et_texte(db_session):
    """Un tableur type une cellule de chiffres tantôt en nombre, tantôt en
    texte — et le texte garde alors les zéros de tête que le nombre perd."""
    n = import_bancaire.normaliser_pour_comparaison
    assert n("00040316718") == n(40316718)
    assert n("0.50") == n(0.5)
    assert n("-45,20") == n(-45.2)
    assert n(" 12 ") == n(12)
    # Ce qui n'est pas exactement un nombre reste du texte : la référence
    # « 1 234 » et le nombre 1234 ne sont pas déclarés identiques sur un
    # découpage de milliers qu'on aurait deviné.
    assert n("1 234") != n(1234)
    assert n("REF 007") != n(7)
    assert n("2026-07-01") != n(2026)
    # Decimal et non float : deux références à vingt chiffres qui ne diffèrent
    # que par le dernier resteraient distinctes.
    assert n("12345678901234567890") != n("12345678901234567891")


def test_normaliser_pour_comparaison_unifie_les_ecritures_dune_date(db_session):
    """Le même jour s'écrit de trois façons selon le format de la colonne et le
    type de fichier ; une seule à l'écran."""
    n = import_bancaire.normaliser_pour_comparaison
    assert n("2026-07-09") == n("2026-07-09T00:00:00")
    assert n("2026-07-09") == n("09/07/2026")
    assert n("2026-07-09") == n(datetime(2026, 7, 9, 0, 0))
    # L'heure est ignorée, comme partout ailleurs dans l'app : une opération
    # n'est datée que du jour.
    assert n("2026-07-09T14:32:00") == n("2026-07-09")
    # Deux jours différents restent deux jours différents.
    assert n("2026-07-09") != n("2026-07-10")
    # Une référence purement numérique n'est pas une date, malgré la forme ISO
    # compacte que datetime.fromisoformat accepterait.
    assert n("20260709") == n(20260709)
    assert n("20260709") != n("2026-07-09")


def test_doublon_detecte_malgre_une_date_ecrite_autrement(db_session):
    """Configuration « uniquement ces colonnes » sur une référence et une date
    — la plus courante, et la plus exposée : la cellule de date sort en texte
    quand la colonne est au format Général, en horodatage quand c'est une vraie
    date Excel. Comparées comme du texte, aucune ligne du fichier n'était plus
    reconnue."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    crud.update_import_preset(
        db_session,
        preset,
        # Colonne 1 (date) et colonne 12 (référence), et rien d'autre.
        colonnes_comparaison=[1, 12],
        mode_comparaison=ModeComparaison.selection.value,
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": get_categorie_id(db_session, "Alimentaire")},
        comptes={"CC Perso": compte.id},
    )
    base = {
        "nature": "Courses",
        "categorie": "Alimentation",
        "montant": -45.2,
        "compte": "CC Perso",
        "reference": "REF-A1",
    }
    # Import initial : la colonne de date porte une vraie date Excel.
    import_bancaire.confirmer(
        db_session,
        preset.id,
        _construire_fichier_avec_reference([{**base, "date": datetime(2026, 7, 9, 0, 0)}]),
        overrides,
    )

    # Réimport du même relevé, dont la colonne de date est cette fois du texte
    # (format Général) — puis à la française, comme le ferait un CSV.
    for date_ecrite in ("2026-07-09", "09/07/2026"):
        preview = import_bancaire.previsualiser(
            db_session,
            preset.id,
            _construire_fichier_avec_reference([{**base, "date": date_ecrite}]),
        )
        assert preview.lignes[0].doublon_de is not None, date_ecrite

    # Un autre jour reste une autre ligne : la date compte toujours.
    preview = import_bancaire.previsualiser(
        db_session,
        preset.id,
        _construire_fichier_avec_reference([{**base, "date": "2026-07-10"}]),
    )
    assert preview.lignes[0].doublon_de is None


def test_doublon_detecte_malgre_un_nombre_ecrit_en_texte(db_session):
    """Deux exports du même compte à quelques minutes d'écart : la référence
    de compte y vaut « 00040316718 » dans l'un et 40316718 dans l'autre, sur la
    seule ligne d'un virement. La colonne n'est pas lue par le preset — rien ne
    le montre donc à l'écran — mais elle entre dans la comparaison en mode
    exclusion, et la ligne cessait d'être reconnue comme doublon."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    base = {
        "date": date(2026, 7, 1),
        "nature": "VIR SEPA M DUPONT",
        "categorie": "Alimentation",
        "montant": -45.2,
        "compte": "CC Perso",
    }
    import_bancaire.confirmer(
        db_session,
        preset.id,
        _construire_fichier_avec_reference([{**base, "reference": "00040316718"}]),
        overrides,
    )

    preview = import_bancaire.previsualiser(
        db_session,
        preset.id,
        _construire_fichier_avec_reference([{**base, "reference": 40316718}]),
    )
    assert preview.lignes[0].doublon_de is not None

    # Une référence réellement différente reste une ligne différente.
    preview = import_bancaire.previsualiser(
        db_session,
        preset.id,
        _construire_fichier_avec_reference([{**base, "reference": 40316719}]),
    )
    assert preview.lignes[0].doublon_de is None


def test_le_type_de_loperation_ne_bloque_pas_la_detection_de_doublon(db_session):
    """La détection compare des LIGNES DE FICHIER, jamais les opérations
    qu'elles ont créées : le type sous lequel une ligne a été importée
    (remboursement de prêt ici) ne doit rien y changer. Les virements SEPA sont
    justement les lignes qu'on reclasse le plus souvent, et celles où un doublon
    raté se remarque le plus."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "VIR SEPA REMB PRET MARIE",
                "categorie": "Alimentation",
                "montant": -200.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": get_categorie_id(db_session, "Alimentaire")},
        comptes={"CC Perso": compte.id},
        lignes={2: schemas.ImportLigneOverride(type_code="remboursement_pret")},
    )
    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)
    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.type_code == "remboursement_pret"

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    assert preview.lignes[0].doublon_de is not None


def test_mappings_categorie_liste_et_suppression(db_session):
    preset = _make_preset(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    crud.set_mapping_categorie(db_session, preset.id, "Alimentation", categorie_id)

    mappings = crud.list_mappings_categorie(db_session, preset.id)
    assert len(mappings) == 1
    assert mappings[0].nom_banque == "Alimentation"
    assert mappings[0].categorie.nom == "Alimentaire"

    assert crud.delete_mapping_categorie(db_session, preset.id, "Alimentation") is True
    assert crud.list_mappings_categorie(db_session, preset.id) == []
    assert crud.delete_mapping_categorie(db_session, preset.id, "Alimentation") is False


def test_mappings_compte_liste_et_suppression(db_session):
    preset = _make_preset(db_session)
    compte = _make_compte(db_session)
    crud.set_mapping_compte(db_session, preset.id, "CC Perso", compte.id)

    mappings = crud.list_mappings_compte(db_session, preset.id)
    assert len(mappings) == 1
    assert mappings[0].nom_banque == "CC Perso"
    assert mappings[0].compte.nom == "CC Perso"

    assert crud.delete_mapping_compte(db_session, preset.id, "CC Perso") is True
    assert crud.list_mappings_compte(db_session, preset.id) == []


def test_mappings_meme_nom_banque_isole_par_preset(db_session):
    preset_a = _make_preset(db_session, nom="Banque A")
    preset_b = _make_preset(db_session, nom="Banque B")
    categorie_alimentaire = get_categorie_id(db_session, "Alimentaire")
    categorie_loisirs = get_categorie_id(db_session, "Loisirs & sorties")

    # Le même nom bancaire "Divers" peut être mappé différemment selon le preset.
    crud.set_mapping_categorie(db_session, preset_a.id, "Divers", categorie_alimentaire)
    crud.set_mapping_categorie(db_session, preset_b.id, "Divers", categorie_loisirs)

    assert crud.get_mapping_categorie(db_session, preset_a.id, "Divers") == categorie_alimentaire
    assert crud.get_mapping_categorie(db_session, preset_b.id, "Divers") == categorie_loisirs


def test_mappings_categorie_tous_presets_porte_la_provenance(db_session):
    """Ce que lit la galerie de la page Règles : les correspondances de TOUS les
    presets, chacune sachant d'où elle vient (compte lié, sinon rien) et à quel
    preset la renvoyer pour la reclasser."""
    compte = _make_compte(db_session, nom="Compte Courant")
    preset_lie = _make_preset(db_session, nom="Banque A")
    crud.update_import_preset(db_session, preset_lie, compte_id=compte.id)
    preset_libre = _make_preset(db_session, nom="Banque B")
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    crud.set_mapping_categorie(db_session, preset_lie.id, "Alimentation", categorie_id)
    crud.set_mapping_categorie(db_session, preset_libre.id, "Alimentation", categorie_id)

    lignes = crud.list_mappings_categorie_tous_presets(db_session)
    par_preset = {m.preset_id: (m, nom_compte) for m, nom_compte in lignes}
    assert set(par_preset) == {preset_lie.id, preset_libre.id}

    mapping_lie, nom_compte_lie = par_preset[preset_lie.id]
    assert mapping_lie.nom_banque == "Alimentation"
    assert nom_compte_lie == "Compte Courant"

    # Preset sans compte lié : présent quand même (outerjoin), sans provenance.
    _, nom_compte_libre = par_preset[preset_libre.id]
    assert nom_compte_libre is None


def test_mappings_compte_tous_presets_fond_les_entrees_identiques(db_session):
    """Comptes et devises s'affichent en une liste commune, sans dire de quel
    preset chacun vient : deux presets qui mappent « CC Perso » vers le même
    compte ne doivent donner qu'UNE ligne, portant les deux presets."""
    from app.routers.import_bancaire import _regrouper_par_cible

    compte = _make_compte(db_session, nom="CC Perso")
    autre_compte = _make_compte(db_session, nom="Livret")
    preset_a = _make_preset(db_session, nom="Banque A")
    preset_b = _make_preset(db_session, nom="Banque B")
    preset_c = _make_preset(db_session, nom="Banque C")
    crud.set_mapping_compte(db_session, preset_a.id, "CC Perso", compte.id)
    crud.set_mapping_compte(db_session, preset_b.id, "CC Perso", compte.id)
    # Même libellé, AUTRE cible : la divergence est justement l'information, la
    # ligne reste distincte.
    crud.set_mapping_compte(db_session, preset_c.id, "CC Perso", autre_compte.id)

    groupes = _regrouper_par_cible(
        crud.list_mappings_compte_tous_presets(db_session), lambda m: m.compte_id
    )
    par_cible = {m.compte_id: sorted(preset_ids) for m, preset_ids in groupes}
    assert par_cible == {
        compte.id: sorted([preset_a.id, preset_b.id]),
        autre_compte.id: [preset_c.id],
    }


def test_import_preset_creation_et_modification(db_session):
    preset = crud.create_import_preset(db_session, "Défaut")
    assert preset.colonnes == COLONNES_IMPORT_PAR_DEFAUT
    assert preset.colonnes_comparaison == []
    assert preset.mode_comparaison == ModeComparaison.exclusion.value
    assert crud.list_import_presets(db_session, DomaineImport.bancaire.value) == [preset]

    nouvelles_colonnes = [
        {"index": 2, "propriete": "date"},
        {"index": 3, "propriete": "nature"},
        {"index": 5, "propriete": "montant"},
    ]
    preset = crud.update_import_preset(db_session, preset, colonnes=nouvelles_colonnes)
    assert preset.colonnes == nouvelles_colonnes
    # Persisté : une nouvelle lecture retrouve la même configuration.
    assert crud.get_import_preset(db_session, preset.id).colonnes == nouvelles_colonnes


def test_import_preset_suppression(db_session):
    preset = crud.create_import_preset(db_session, "Banque X")
    crud.delete_import_preset(db_session, preset)
    assert crud.get_import_preset(db_session, preset.id) is None


def _fichier_trois_colonnes(lignes: list[list]) -> bytes:
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    for ligne in lignes:
        feuille.append(ligne)
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


_COLONNES_TROIS = [
    {"index": 1, "propriete": "date"},
    {"index": 2, "propriete": "nature"},
    {"index": 3, "propriete": "montant"},
]


def test_lire_lignes_brutes_lit_la_premiere_ligne_par_defaut(db_session):
    """Défaut applicatif : pas d'en-tête supposé. Tous les formats n'en ont
    pas, et le sauter systématiquement perdait une opération par import."""
    contenu = _fichier_trois_colonnes(
        [
            [date(2026, 7, 1), "Courses", -45.2],
            [date(2026, 7, 2), "Essence", -60.0],
        ]
    )

    lignes = import_bancaire.lire_lignes_brutes(contenu, _COLONNES_TROIS)

    assert [l["nature"] for l in lignes] == ["Courses", "Essence"]
    # Numéro de ligne physique dans le fichier, 1-based (ce que voit l'utilisateur).
    assert [l["ligne"] for l in lignes] == [1, 2]


def test_lire_lignes_brutes_saute_len_tete_quand_le_preset_le_demande(db_session):
    contenu = _fichier_trois_colonnes(
        [
            ["Date", "Libellé", "Montant"],
            [date(2026, 7, 1), "Courses", -45.2],
        ]
    )

    lignes = import_bancaire.lire_lignes_brutes(
        contenu, _COLONNES_TROIS, ignorer_premiere_ligne=True
    )

    assert [l["nature"] for l in lignes] == ["Courses"]
    assert lignes[0]["ligne"] == 2


def test_detecter_delimiteur_prefere_le_point_virgule_malgre_les_virgules_decimales(db_session):
    """Régression : csv.Sniffer élisait "," sur un relevé français sans
    en-tête, parce que les montants ("-45,2") contiennent des virgules. Le
    fichier était alors découpé n'importe comment (nature="11", montant=None)."""
    sans_entete = "01/07/2026;;;Courses;;Alim;-45,2;;;CC;;\r\n02/07/2026;;;Essence;;Transp;-60,5;;;CC;;\r\n"
    assert import_bancaire._detecter_delimiteur(sans_entete) == ";"

    # Les formats non ambigus restent correctement détectés.
    assert (
        import_bancaire._detecter_delimiteur(
            "date,nature,montant\r\n2026-07-01,Courses,-45.20\r\n2026-07-02,Essence,-60.50\r\n"
        )
        == ","
    )
    assert (
        import_bancaire._detecter_delimiteur("date\tnature\tmontant\r\n2026-07-01\tCourses\t-45.20\r\n")
        == "\t"
    )
    # Échantillon inexploitable : repli sur le standard français.
    assert import_bancaire._detecter_delimiteur("") == ";"


def test_lire_lignes_brutes_csv_francais_sans_entete(db_session):
    contenu = (
        "01/07/2026;;;Courses Monoprix;;Alimentation;-45,2;;;CC Perso;;\r\n"
        "02/07/2026;;;Essence Total;;Transport;-60,5;;;CC Perso;;\r\n"
    ).encode("utf-8")

    lignes = import_bancaire.lire_lignes_brutes(contenu, COLONNES_IMPORT_PAR_DEFAUT)

    assert [l["nature"] for l in lignes] == ["Courses Monoprix", "Essence Total"]
    assert import_bancaire.parser_montant(lignes[0]["montant_brut"]) == -45.2
    assert [l["compte_banque"] for l in lignes] == ["CC Perso", "CC Perso"]


def test_apercu_fichier_expose_les_lignes_brutes_et_le_mapping_des_colonnes(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session, ignorer_premiere_ligne=True)
    contenu = _construire_fichier_avec_reference(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
                "reference": "REF-001",
            }
        ]
    )

    apercu = import_bancaire.previsualiser(
        db_session, preset.id, contenu, compte_id_defaut=compte.id
    ).apercu_fichier

    # En-tête + ligne de données : les deux sont exposées telles quelles, y
    # compris l'en-tête ignoré (affiché barré côté frontend).
    assert apercu.total_lignes == 2
    assert apercu.premiere_ligne_ignoree is True
    assert apercu.lignes[0][0] == "Colonne 1"
    assert apercu.lignes[1][3] == "Courses"
    # Colonne 12 lue dans le fichier mais non mappée : présente dans les
    # lignes, absente du mapping (affichée en gris).
    assert apercu.lignes[1][11] == "REF-001"
    assert apercu.proprietes_par_colonne == {
        "1": "date",
        "4": "nature",
        "6": "categorie_banque",
        "7": "montant",
        "10": "compte_banque",
    }


def test_apercu_fichier_expose_toutes_les_lignes_dun_gros_fichier(db_session):
    """Plus aucune troncature côté serveur : une colonne décalée ne l'est pas
    forcément dès les premières lignes, et le vérifier demande de pouvoir tout
    parcourir. C'est le frontend qui borne la hauteur affichée."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {"date": date(2026, 7, 1), "nature": f"Op {i}", "montant": -1.0, "compte": "CC Perso"}
            for i in range(80)
        ]
    )

    apercu = import_bancaire.previsualiser(
        db_session, preset.id, contenu, compte_id_defaut=compte.id
    ).apercu_fichier

    # 80 lignes de données + la ligne d'en-tête écrite par _construire_fichier :
    # l'aperçu montre le fichier tel quel, en-tête compris.
    assert apercu.total_lignes == 81
    assert len(apercu.lignes) == 81
    assert apercu.lignes[-1][3] == "Op 79"


def test_previsualiser_respecte_ignorer_premiere_ligne_du_preset(db_session):
    compte = _make_compte(db_session)
    contenu = _fichier_trois_colonnes(
        [
            [date(2026, 7, 1), "Courses", -45.2],
            [date(2026, 7, 2), "Essence", -60.0],
        ]
    )

    preset_sans_entete = _make_preset(
        db_session, colonnes=_COLONNES_TROIS, ignorer_premiere_ligne=False
    )
    preview = import_bancaire.previsualiser(
        db_session, preset_sans_entete.id, contenu, compte_id_defaut=compte.id
    )
    assert [l.nature for l in preview.lignes] == ["Courses", "Essence"]

    preset_avec_entete = _make_preset(
        db_session, nom="Avec en-tête", colonnes=_COLONNES_TROIS, ignorer_premiere_ligne=True
    )
    preview = import_bancaire.previsualiser(
        db_session, preset_avec_entete.id, contenu, compte_id_defaut=compte.id
    )
    assert [l.nature for l in preview.lignes] == ["Essence"]


def test_import_avec_colonnes_personnalisees(db_session):
    compte = _make_compte(db_session)
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    # Format à 5 colonnes, sans catégorie bancaire ni compte : tout va dans
    # "Autres" et utilise le compte par défaut fourni à la confirmation.
    preset = _make_preset(
        db_session,
        colonnes=[
            {"index": 1, "propriete": "date"},
            {"index": 2, "propriete": "nature"},
            {"index": 3, "propriete": "montant"},
        ],
    )
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.append(["Date", "Libellé", "Montant"])
    feuille.append([date(2026, 7, 1), "Courses", -45.2])
    tampon = io.BytesIO()
    classeur.save(tampon)
    contenu = tampon.getvalue()

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu, compte_id_defaut=compte.id)
    assert preview.lignes[0].nature == "Courses"
    assert preview.lignes[0].montant == 45.2
    assert preview.lignes[0].compte_id == compte.id
    assert preview.lignes[0].categorie_suggestion_auto is True

    overrides = schemas.ImportMappingOverrides()
    resultat = import_bancaire.confirmer(
        db_session, preset.id, contenu, overrides, compte_id_defaut=compte.id
    )
    assert resultat.operations_creees == 1
    operation = db_session.query(models.Operation).one()
    assert operation.compte_id == compte.id
    assert operation.categorie_id == get_categorie_id(db_session, "Autres")


def test_delete_all_operations(db_session):
    compte = _make_compte(db_session)
    for i in range(3):
        crud.create_operation(
            db_session,
            schemas.OperationCreate(
                date=date(2026, 7, 1 + i),
                compte_id=compte.id,
                monnaie_id=get_monnaie_id(db_session),
                type_id=get_type_id(db_session, "classique"),
                categorie_id=get_categorie_id(db_session, "Alimentaire"),
                nature=f"Opération {i}",
                montant=10.0,
                statut=Statut.reel,
            ),
        )

    nb = crud.delete_all_operations(db_session)

    assert nb == 3
    assert db_session.query(models.Operation).count() == 0


# ---------- Preset lié à un compte bancaire ----------


def test_preset_lie_affecte_toutes_les_lignes_a_son_compte(db_session):
    """Le relevé d'un compte précis ne nomme nulle part le compte concerné :
    lié une fois, le preset le pose sur chaque ligne, sans mapping ni compte
    choisi pour le fichier."""
    compte = _make_compte(db_session, "CC Perso")
    preset = _make_preset(
        db_session,
        colonnes=[
            {"index": 1, "propriete": "date"},
            {"index": 4, "propriete": "nature"},
            {"index": 7, "propriete": "montant"},
        ],
    )
    crud.update_import_preset(db_session, preset, compte_id=compte.id)

    contenu = _construire_fichier(
        [
            {"date": date(2026, 7, 1), "nature": "Courses", "montant": -45.2},
            {"date": date(2026, 7, 2), "nature": "Salaire", "montant": 1500.0},
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    assert [l.compte_id for l in preview.lignes] == [compte.id, compte.id]
    # Plus rien à faire correspondre : c'est tout l'intérêt de la liaison.
    assert preview.comptes_inconnus == []


def test_preset_lie_prime_sur_la_colonne_compte_du_fichier(db_session):
    """Deux comptes ne peuvent pas sortir d'un relevé qui appartient à un seul :
    le choix explicite de l'utilisateur passe devant ce que dit le fichier, et
    devant les correspondances mémorisées."""
    compte_lie = _make_compte(db_session, "CC Perso")
    autre = creer_compte(db_session, "Livret A", type_nom="épargne")
    preset = _make_preset(db_session)
    crud.set_mapping_compte(db_session, preset.id, "Compte Bis", autre.id)
    crud.update_import_preset(db_session, preset, compte_id=compte_lie.id)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "Compte Bis",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    assert preview.lignes[0].compte_id == compte_lie.id
    # Le libellé du fichier reste affiché tel quel : il n'a simplement plus voix
    # au chapitre pour désigner le compte.
    assert preview.lignes[0].nom_banque_compte == "Compte Bis"


def test_preset_lie_oriente_le_virement_selon_le_signe_du_montant(db_session):
    """Le compte lié joue exactement le rôle d'un compte résolu par colonne :
    émetteur si le montant bancaire est négatif, récepteur s'il est positif."""
    compte_lie = _make_compte(db_session, "CC Perso")
    destination = creer_compte(db_session, "Livret A", type_nom="épargne")
    preset = _make_preset(
        db_session,
        colonnes=[
            {"index": 1, "propriete": "date"},
            {"index": 4, "propriete": "nature"},
            {"index": 7, "propriete": "montant"},
        ],
    )
    crud.update_import_preset(db_session, preset, compte_id=compte_lie.id)

    contenu = _construire_fichier(
        [{"date": date(2026, 7, 1), "nature": "Vers Livret A", "montant": -100.0}]
    )
    overrides = schemas.ImportMappingOverrides(
        lignes={
            2: schemas.ImportLigneOverride(
                type_code="virement", compte_id_autre=destination.id
            )
        }
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 2
    sortante = (
        db_session.query(models.Operation)
        .filter(models.Operation.sens == Sens.transfert_sortant)
        .one()
    )
    entrante = (
        db_session.query(models.Operation)
        .filter(models.Operation.sens == Sens.transfert_entrant)
        .one()
    )
    # Montant négatif : le compte lié est l'émetteur.
    assert sortante.compte_id == compte_lie.id
    assert entrante.compte_id == destination.id


def test_preset_lie_recepteur_quand_le_montant_est_positif(db_session):
    compte_lie = _make_compte(db_session, "CC Perso")
    source = creer_compte(db_session, "Livret A", type_nom="épargne")
    preset = _make_preset(
        db_session,
        colonnes=[
            {"index": 1, "propriete": "date"},
            {"index": 4, "propriete": "nature"},
            {"index": 7, "propriete": "montant"},
        ],
    )
    crud.update_import_preset(db_session, preset, compte_id=compte_lie.id)

    contenu = _construire_fichier(
        [{"date": date(2026, 7, 1), "nature": "Depuis Livret A", "montant": 100.0}]
    )
    overrides = schemas.ImportMappingOverrides(
        lignes={
            2: schemas.ImportLigneOverride(type_code="virement", compte_id_autre=source.id)
        }
    )

    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)

    assert resultat.operations_creees == 2
    sortante = (
        db_session.query(models.Operation)
        .filter(models.Operation.sens == Sens.transfert_sortant)
        .one()
    )
    entrante = (
        db_session.query(models.Operation)
        .filter(models.Operation.sens == Sens.transfert_entrant)
        .one()
    )
    assert entrante.compte_id == compte_lie.id
    assert sortante.compte_id == source.id


def test_preset_non_lie_garde_la_resolution_par_le_fichier(db_session):
    """La liaison est opt-in : sans elle, rien ne change."""
    compte = _make_compte(db_session, "CC Perso")
    preset = _make_preset(db_session)
    crud.set_mapping_compte(db_session, preset.id, "CC Perso", compte.id)

    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            }
        ]
    )

    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)

    assert preset.compte_id is None
    assert preview.lignes[0].compte_id == compte.id


def test_delier_un_preset_le_remet_dans_l_etat_d_avant(db_session):
    """compte_id=None doit vouloir dire « délie », pas « ne touche à rien » --
    sans quoi une liaison serait impossible à défaire."""
    compte = _make_compte(db_session, "CC Perso")
    preset = _make_preset(db_session)
    crud.update_import_preset(db_session, preset, compte_id=compte.id)

    crud.update_import_preset(db_session, preset, compte_id=None)

    assert preset.compte_id is None


# ---------- Réglages de lecture en dernier recours (délimiteur, séparateur
# décimal), proposés par le frontend quand l'aperçu ne comprend plus le
# fichier (cf. previsualiser/confirmer et le module parser_montant).


def test_parser_montant_par_defaut_reste_permissif():
    """Comportement historique inchangé quand separateur_decimal n'est pas
    précisé : virgule française lue en décimale, un point déjà présent laissé
    tel quel."""
    assert import_bancaire.parser_montant("-45,2") == -45.2
    assert import_bancaire.parser_montant("45.2") == 45.2


def test_parser_montant_avec_separateur_decimal_point_lit_les_milliers_en_virgule():
    """Format anglo-saxon : la virgule sépare les milliers, le point est la
    décimale. Le mode permissif par défaut s'y tromperait (il convertirait la
    virgule en point)."""
    assert import_bancaire.parser_montant("1,234.56", separateur_decimal=".") == 1234.56
    assert import_bancaire.parser_montant("1,234", separateur_decimal=".") == 1234.0


def test_parser_montant_avec_separateur_decimal_virgule_lit_les_milliers_en_point():
    """Format européen non français : le point sépare les milliers, la
    virgule est la décimale. Le mode permissif par défaut s'y tromperait (le
    point resterait, cassant le nombre)."""
    assert import_bancaire.parser_montant("1.234,56", separateur_decimal=",") == 1234.56


def test_lire_lignes_brutes_avec_un_delimiteur_impose(db_session):
    """Un délimiteur hors des trois candidats de la détection automatique
    (« ; », tabulation, « , ») n'est lu correctement qu'en le précisant."""
    contenu = (
        "|".join(f"Colonne {i}" for i in range(1, 13)) + "\r\n"
        "01/07/2026|||Courses||Alimentation|-45,2|||CC Perso||\r\n"
    ).encode("utf-8")

    # Sans le préciser : la détection automatique ne reconnaît pas "|", et
    # relit tout sur une seule colonne.
    sans_delimiteur = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True
    )
    assert sans_delimiteur[0]["nature"] is None

    avec_delimiteur = import_bancaire.lire_lignes_brutes(
        contenu, COLONNES_IMPORT_PAR_DEFAUT, ignorer_premiere_ligne=True, delimiteur="|"
    )
    assert avec_delimiteur[0]["nature"] == "Courses"
    assert avec_delimiteur[0]["montant_brut"] == "-45,2"


def test_previsualiser_avec_delimiteur_et_separateur_decimal_impose(db_session):
    """Bout en bout : un relevé pipe-délimité en format anglo-saxon (virgule
    de milliers, point décimal) reste illisible par défaut et se résout une
    fois les deux réglages précisés — le scénario que le frontend propose
    quand l'aperçu détecte une majorité de lignes en erreur."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    crud.set_mapping_compte(db_session, preset.id, "CC Perso", compte.id)

    contenu = (
        "|".join(f"Colonne {i}" for i in range(1, 13)) + "\r\n"
        "01/07/2026|||Courses||Alimentation|-1,234.56|||CC Perso||\r\n"
    ).encode("utf-8")

    rate = import_bancaire.previsualiser(db_session, preset.id, contenu)
    assert rate.lignes[0].erreur is not None
    assert "illisible" in rate.lignes[0].erreur

    reussi = import_bancaire.previsualiser(
        db_session, preset.id, contenu, delimiteur="|", separateur_decimal="."
    )
    assert reussi.lignes[0].erreur is None
    assert reussi.lignes[0].montant == 1234.56
    assert reussi.lignes[0].compte_id == compte.id


# ---------- Annulation d'un import (retour en arrière depuis l'historique) ----------


def _importer_deux_lignes(db_session, compte, preset):
    """Un import banal de deux opérations classiques, point de départ des tests
    d'annulation ci-dessous."""
    categorie_id = get_categorie_id(db_session, "Alimentaire")
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Courses",
                "categorie": "Alimentation",
                "montant": -45.2,
                "compte": "CC Perso",
            },
            {
                "date": date(2026, 7, 2),
                "nature": "Essence",
                "categorie": "Alimentation",
                "montant": -60.0,
                "compte": "CC Perso",
            },
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        categories={"Alimentation": categorie_id}, comptes={"CC Perso": compte.id}
    )
    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)
    return contenu, resultat


def test_annuler_un_import_supprime_ses_operations_et_sa_trace(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, resultat = _importer_deux_lignes(db_session, compte, preset)
    assert resultat.operations_creees == 2
    assert db_session.query(models.Operation).count() == 2

    annulation = import_bancaire.annuler_import(db_session, resultat.historique_id)

    assert annulation.operations_supprimees == 2
    assert annulation.historique_supprime is True
    assert db_session.query(models.Operation).count() == 0
    assert crud.get_import_historique(db_session, preset.id) == []


def test_annuler_un_import_vide_le_stock_anti_doublons(db_session):
    """Le point crucial : sans ça, le relevé resterait « déjà importé » alors
    qu'il n'en reste rien en base, et le réimporter le verrait comme un doublon
    de lignes disparues."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu, resultat = _importer_deux_lignes(db_session, compte, preset)
    assert len(crud.list_lignes_import_brutes(db_session, preset.id)) == 2

    import_bancaire.annuler_import(db_session, resultat.historique_id)

    assert crud.list_lignes_import_brutes(db_session, preset.id) == []
    # Et le même fichier repasse comme neuf, sans aucun doublon signalé.
    preview = import_bancaire.previsualiser(db_session, preset.id, contenu)
    assert all(ligne.doublon_de is None for ligne in preview.lignes)


def test_annuler_un_import_supprime_les_deux_jambes_dun_virement(db_session):
    """Le stock ne retient que la jambe sortante : sans le rattrapage par
    virement_id, l'annulation laisserait une demi-écriture sur le compte
    d'en face."""
    compte_source = _make_compte(db_session, "CC Perso")
    compte_destination = _make_compte(db_session, "Livret A")
    preset = _make_preset(db_session)
    contenu = _construire_fichier(
        [
            {
                "date": date(2026, 7, 1),
                "nature": "Vers Livret A",
                "categorie": "Divers",
                "montant": -100.0,
                "compte": "CC Perso",
            }
        ]
    )
    overrides = schemas.ImportMappingOverrides(
        comptes={"CC Perso": compte_source.id},
        lignes={
            2: schemas.ImportLigneOverride(
                type_code="virement", compte_id_autre=compte_destination.id
            )
        },
    )
    resultat = import_bancaire.confirmer(db_session, preset.id, contenu, overrides)
    assert db_session.query(models.Operation).count() == 2

    annulation = import_bancaire.annuler_import(db_session, resultat.historique_id)

    assert annulation.operations_supprimees == 2
    assert db_session.query(models.Operation).count() == 0


def test_annuler_un_import_ignore_une_operation_deja_supprimee_a_la_main(db_session):
    """Le CASCADE a déjà retiré sa ligne du stock : elle ne se compte pas une
    seconde fois, et l'annulation ne trébuche pas dessus."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, resultat = _importer_deux_lignes(db_session, compte, preset)
    premiere = db_session.query(models.Operation).first()
    crud.delete_operation(db_session, premiere)

    annulation = import_bancaire.annuler_import(db_session, resultat.historique_id)

    assert annulation.operations_supprimees == 1
    assert db_session.query(models.Operation).count() == 0


def test_annuler_un_import_ne_touche_pas_aux_autres_imports(db_session):
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, premier = _importer_deux_lignes(db_session, compte, preset)

    autre_contenu = _construire_fichier(
        [
            {
                "date": date(2026, 8, 15),
                "nature": "Restaurant",
                "categorie": "Alimentation",
                "montant": -30.0,
                "compte": "CC Perso",
            }
        ]
    )
    second = import_bancaire.confirmer(
        db_session,
        preset.id,
        autre_contenu,
        schemas.ImportMappingOverrides(comptes={"CC Perso": compte.id}),
    )
    assert db_session.query(models.Operation).count() == 3

    import_bancaire.annuler_import(db_session, premier.historique_id)

    restantes = db_session.query(models.Operation).all()
    assert len(restantes) == 1
    assert restantes[0].nature == "Restaurant"
    historique = crud.get_import_historique(db_session, preset.id)
    assert [h.id for h in historique] == [second.historique_id]


def test_annuler_un_import_inexistant_ne_fait_rien(db_session):
    annulation = import_bancaire.annuler_import(db_session, 9999)

    assert annulation.operations_supprimees == 0
    assert annulation.historique_supprime is False


def test_compter_operations_annulables_suit_les_suppressions_manuelles(db_session):
    """Ce que le frontend affiche avant de proposer d'annuler : ce qui reste
    aujourd'hui, pas ce que l'import avait créé le jour même."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, resultat = _importer_deux_lignes(db_session, compte, preset)
    assert crud.compter_operations_annulables(db_session, preset.id) == {
        resultat.historique_id: {"annulables": 2, "sans_lien": 0}
    }

    crud.delete_operation(db_session, db_session.query(models.Operation).first())

    assert crud.compter_operations_annulables(db_session, preset.id) == {
        resultat.historique_id: {"annulables": 1, "sans_lien": 0}
    }


def test_une_ligne_de_stock_sans_operation_est_comptee_a_part(db_session):
    """Les lignes d'avant la migration 0016 n'ont jamais désigné leur
    opération : leur import ne sera JAMAIS annulable, ce qui n'est pas la même
    chose qu'un import dont on a tout supprimé — d'où le compteur séparé, qui
    permet à l'écran de le dire."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, resultat = _importer_deux_lignes(db_session, compte, preset)
    # Reproduit une ligne historique : rattachée à l'import, sans opération.
    crud.create_ligne_import_brute(
        db_session,
        preset_id=preset.id,
        donnees={"1": "ancienne"},
        import_historique_id=resultat.historique_id,
        operation_id=None,
    )

    assert crud.compter_operations_annulables(db_session, preset.id) == {
        resultat.historique_id: {"annulables": 2, "sans_lien": 1}
    }


def test_la_route_historique_expose_ce_qui_est_annulable_et_pourquoi(db_session):
    """Ce que la page Import lit pour décider d'afficher le bouton, ou le
    message qui explique son absence."""
    from app.routers.import_bancaire import get_historique

    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    _, resultat = _importer_deux_lignes(db_session, compte, preset)

    [lecture] = get_historique(preset.id, db=db_session)
    assert lecture.operations_annulables == 2
    assert lecture.raison_non_annulable is None

    # Tout supprimé à la main : le bouton disparaît, et l'écran doit dire
    # « plus rien », pas « import trop ancien ».
    for operation in db_session.query(models.Operation).all():
        crud.delete_operation(db_session, operation)

    [lecture] = get_historique(preset.id, db=db_session)
    assert lecture.operations_annulables == 0
    assert lecture.raison_non_annulable == "deja_supprime"


def test_la_route_dannulation_refuse_un_import_dun_autre_preset(db_session):
    """Scopée au preset comme le reste de la page : sans ce contrôle, l'écran
    d'une banque pourrait annuler en silence l'import d'une autre."""
    from fastapi import HTTPException

    from app.routers.import_bancaire import annuler_import as route_annuler

    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    autre_preset = _make_preset(db_session, nom="Autre banque")
    _, resultat = _importer_deux_lignes(db_session, compte, preset)

    with pytest.raises(HTTPException) as erreur:
        route_annuler(autre_preset.id, resultat.historique_id, db=db_session)

    assert erreur.value.status_code == 404
    # Et rien n'a bougé.
    assert db_session.query(models.Operation).count() == 2


def test_un_reglement_lie_est_rattache_a_son_import(db_session):
    """Les règlements liés naissent APRÈS le confirm groupé, un par un : sans
    ce rattachement, eux seuls survivraient à l'annulation de leur import."""
    compte = _make_compte(db_session)
    preset = _make_preset(db_session)
    contenu, resultat = _importer_deux_lignes(db_session, compte, preset)

    # Une opération créée hors du confirm, puis déclarée sur la ligne 2.
    operation = crud.create_operation_importee(
        db_session,
        date_operation=date(2026, 7, 3),
        compte_id=compte.id,
        type_id=get_type_id(db_session, "remboursements"),
        categorie_id=None,
        nature="Remboursement lié",
        montant=15.0,
        monnaie_id=compte.monnaie_principale_id,
        montant_du=0.0,
        sens=None,
        statut=Statut.reel,
    )
    trouvee = import_bancaire.enregistrer_ligne_brute(
        db_session,
        preset.id,
        contenu,
        2,
        operation.id,
        import_historique_id=resultat.historique_id,
    )
    assert trouvee is True

    annulation = import_bancaire.annuler_import(db_session, resultat.historique_id)

    assert annulation.operations_supprimees == 3
    assert db_session.query(models.Operation).count() == 0
